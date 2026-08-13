"""本文件把全工站MES工作簿解析为产品、工站事件和参数明细。"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
import re

import pandas as pd
from openpyxl import load_workbook

from .station_sources import StationCatalog


COMMON_ALIASES = {
    "dmc_raw": {"identno", "uniquepartidentno"},
    "test_date": {"testdate", "packagingdate"},
    "state_raw": {"state", "statetest"},
    "line": {"line"}, "st": {"st"}, "si": {"si"}, "fu": {"fu"},
    "wp": {"wp"}, "tp": {"tp"}, "ttnr": {"ttnr"},
    "variant": {"variant"}, "batch": {"batch"},
}


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_state(value: Any) -> str:
    text = _text(value).upper()
    if text == "OK": return "OK"
    if text == "NOK": return "NOK"
    if "SCRAP" in text: return "SCRAPPED"
    if text.startswith("OTHER"): return "OTHERS"
    return "UNKNOWN"


@dataclass
class StationWorkbookData:
    products: pd.DataFrame
    events: pd.DataFrame
    parameters: pd.DataFrame
    package: pd.DataFrame
    warnings: list[str] = field(default_factory=list)
    source_sheets: tuple[str, ...] = ()

    def station_events(self, station_id: str) -> pd.DataFrame:
        return self.events[self.events["station_id"].eq(station_id)].copy()


def _headers(row: tuple[Any, ...]) -> tuple[list[str], list[str]]:
    originals = [_text(value) for value in row]
    counts: dict[str, int] = {}
    unique: list[str] = []
    for index, original in enumerate(originals, 1):
        base = original or f"Unnamed_{index}"
        counts[base] = counts.get(base, 0) + 1
        unique.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return unique, originals


def _block_frames(worksheet) -> list[tuple[str, pd.DataFrame, list[str]]]:
    """Read either one normal table or vertically stacked station sample tables."""
    rows = list(worksheet.iter_rows(values_only=True))
    starts = [
        index for index, row in enumerate(rows)
        if row and any(_norm(value) in {"identno", "uniquepartidentno"} for value in row)
    ]
    blocks: list[tuple[str, pd.DataFrame, list[str]]] = []
    for position, start in enumerate(starts):
        end = starts[position + 1] - 1 if position + 1 < len(starts) else len(rows)
        title = worksheet.title
        if start and sum(value not in (None, "") for value in rows[start - 1]) == 1:
            title = _text(rows[start - 1][0]) or title
        columns, originals = _headers(rows[start])
        values = [row for row in rows[start + 1:end] if row and row[0] not in (None, "")]
        if values:
            blocks.append((title, pd.DataFrame(values, columns=columns), originals))
    return blocks


def _canonical_columns(columns: list[str], originals: list[str]) -> dict[str, str]:
    lookup = {_norm(original): unique for unique, original in zip(columns, originals)}
    result: dict[str, str] = {}
    for canonical, aliases in COMMON_ALIASES.items():
        for alias in aliases:
            if alias in lookup:
                result[canonical] = lookup[alias]
                break
    return result


def _station_id(row: pd.Series, mapping: dict[str, str], catalog: StationCatalog) -> str:
    required = ("line", "st", "si", "fu", "wp")
    if not all(name in mapping for name in required):
        return ""
    parts = [_text(row[mapping[name]]) for name in required]
    location = ".".join(parts)
    station = catalog.station_for_location(location)
    if station:
        return station.id
    # Some MES exports contain an inconsistent SI for VI. Line+ST+WP still uniquely
    # identifies the configured station, so use it as an explicit, auditable fallback.
    candidates = []
    for item in catalog.stations:
        location_parts = item.location.split(".")
        if len(location_parts) >= 5 and (
            location_parts[0] == parts[0]
            and location_parts[1] == parts[1]
            and location_parts[4] == parts[4]
        ):
            candidates.append(item)
    return candidates[0].id if len(candidates) == 1 else ""


def _timestamp(value: Any) -> pd.Timestamp:
    text = _text(value)
    if not text:
        return pd.NaT
    if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", text):
        return pd.to_datetime(text, errors="coerce", yearfirst=True)
    return pd.to_datetime(text, errors="coerce", dayfirst=True)


def load_station_workbook(path: Path, catalog: StationCatalog) -> StationWorkbookData:
    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"Excel工作簿不存在：{resolved}")
    workbook = load_workbook(
        resolved, read_only=True, data_only=True,
        keep_vba=resolved.suffix.lower() == ".xlsm",
    )
    event_rows: list[dict[str, Any]] = []
    parameter_rows: list[dict[str, Any]] = []
    package_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    source_sheets = tuple(workbook.sheetnames)
    try:
        for sheet_name in workbook.sheetnames:
            blocks = _block_frames(workbook[sheet_name])
            if not blocks:
                warnings.append(f"工作表 {sheet_name} 未找到Ident No.表头，已跳过")
                continue
            for block_title, frame, originals in blocks:
                mapping = _canonical_columns(list(frame.columns), originals)
                if "dmc_raw" not in mapping:
                    continue
                is_package = any(_norm(value) == "uniquepartidentno" for value in originals)
                for row_number, (_, row) in enumerate(frame.iterrows(), 2):
                    dmc = _text(row[mapping["dmc_raw"]])
                    if not dmc:
                        continue
                    if is_package:
                        package_rows.append({
                            "dmc_raw": dmc, "source_sheet": sheet_name,
                            "source_row": row_number,
                            **{original: row[column] for column, original in zip(frame.columns, originals) if original},
                        })
                        continue
                    station_id = _station_id(row, mapping, catalog)
                    timestamp = _timestamp(
                        row[mapping["test_date"]] if "test_date" in mapping else None
                    )
                    state_raw = row[mapping["state_raw"]] if "state_raw" in mapping else ""
                    event_id = f"{sheet_name}:{block_title}:{row_number}"
                    event = {
                        "event_id": event_id, "dmc_raw": dmc, "station_id": station_id,
                        "station_title": block_title, "test_date": timestamp,
                        "state_raw": _text(state_raw), "state": normalize_state(state_raw),
                        "source_sheet": sheet_name, "source_row": row_number,
                    }
                    for canonical in ("line", "st", "si", "fu", "wp", "tp", "ttnr", "variant", "batch"):
                        event[canonical] = row[mapping[canonical]] if canonical in mapping else None
                    # Keep result labels needed by AOI/VI truth review.
                    for column, original in zip(frame.columns, originals):
                        normalized = _norm(original)
                        if any(token in normalized for token in (
                            "failurecode", "failurescode", "blockcode", "resultfrom",
                            "overallresultstate", "forwardresult", "backwardresult",
                        )):
                            event[original] = row[column]
                    if not station_id:
                        warnings.append(
                            f"{sheet_name}第{row_number}行无法由Line/ST/SI/FU/WP映射工站"
                        )
                    event_rows.append(event)
                    common_sources = set(mapping.values())
                    for index, (column, original) in enumerate(zip(frame.columns, originals)):
                        if not original or column in common_sources or _norm(original).startswith("tolerance"):
                            continue
                        tolerance_column = None
                        if index + 1 < len(originals) and _norm(originals[index + 1]).startswith("tolerance"):
                            tolerance_column = frame.columns[index + 1]
                        raw_value = row[column]
                        if pd.isna(raw_value) or (isinstance(raw_value, str) and not raw_value.strip()):
                            continue
                        numeric = pd.to_numeric(pd.Series([raw_value]), errors="coerce").iloc[0]
                        parameter_rows.append({
                            "event_id": event_id, "dmc_raw": dmc, "station_id": station_id,
                            "test_date": timestamp, "parameter_name": original,
                            "raw_value": raw_value,
                            "numeric_value": numeric if pd.notna(numeric) else pd.NA,
                            "tolerance_raw": row[tolerance_column] if tolerance_column else None,
                            "parameter_type": "numeric" if pd.notna(numeric) else "categorical",
                            "source_sheet": sheet_name,
                        })
    finally:
        workbook.close()

    events = pd.DataFrame(event_rows)
    parameters = pd.DataFrame(parameter_rows)
    package = pd.DataFrame(package_rows)
    if events.empty:
        raise ValueError("工作簿中没有可识别的工站记录")
    if not parameters.empty:
        parameter_counts = parameters.groupby("event_id").size()
        events["_parameter_count"] = events["event_id"].map(parameter_counts).fillna(0)
        has_detailed_aoi = events.groupby(["dmc_raw", "station_id"])["_parameter_count"].transform("max").gt(0)
        redundant_aoi_history = (
            events["station_id"].astype(str).str.endswith("_aoi")
            & has_detailed_aoi & events["_parameter_count"].eq(0)
        )
        events = events[~redundant_aoi_history]
        events = (
            events.sort_values("_parameter_count", kind="stable")
            .drop_duplicates(["dmc_raw", "station_id", "test_date"], keep="last")
            .drop(columns="_parameter_count")
        )
    events = events.sort_values(["test_date", "dmc_raw"], na_position="last", kind="stable").reset_index(drop=True)
    product_columns = ["dmc_raw", "ttnr", "variant", "batch"]
    products = (
        events.sort_values("test_date", na_position="last")
        .groupby("dmc_raw", as_index=False)[product_columns[1:]].first()
    )
    return StationWorkbookData(
        products=products, events=events, parameters=parameters, package=package,
        warnings=list(dict.fromkeys(warnings)), source_sheets=source_sheets,
    )


def enrich_products_with_station_truth(
    products: pd.DataFrame, workbook: StationWorkbookData, station_id: str
) -> pd.DataFrame:
    """Attach the selected AOI truth and order images by its real MES timestamp."""
    result = products.copy()
    history = workbook.station_events(station_id)
    if history.empty:
        result["truth_match"] = "unmatched"
        return result
    history = history.sort_values("test_date", na_position="last", kind="stable")
    selected_rows: list[pd.Series] = []
    for _, product in result.iterrows():
        candidates = history[history["dmc_raw"].astype(str).eq(str(product["dmc_raw"]))]
        capture = pd.to_datetime(product.get("capture_date"), errors="coerce")
        if pd.notna(capture) and not candidates.empty:
            same_day = candidates[
                pd.to_datetime(candidates["test_date"], errors="coerce").dt.date.eq(capture.date())
            ]
            if not same_day.empty:
                candidates = same_day
        if not candidates.empty:
            selected_rows.append(candidates.iloc[-1])
    latest = pd.DataFrame(selected_rows).drop_duplicates("dmc_raw", keep="last")
    truth_columns = [
        column for column in latest.columns
        if column in {"dmc_raw", "test_date", "state", "state_raw", "event_id"}
        or any(token in _norm(column) for token in ("failurecode", "overallresultstate"))
    ]
    latest = latest[truth_columns].rename(columns={
        "test_date": "aoi_test_date", "state": "aoi_state",
        "state_raw": "aoi_state_raw", "event_id": "aoi_event_id",
    })
    result = result.merge(latest, on="dmc_raw", how="left")
    result["truth_match"] = result["aoi_event_id"].notna().map({True: "matched", False: "unmatched"})
    vi_station_id = {
        "35_5s_aoi": "35_5s_vi", "57_5x_aoi": "57_5x_vi",
        "conveyor_7s_aoi": "conveyor_7s_vi", "conveyor_7x_aoi": "conveyor_7x_vi",
    }.get(station_id)
    if vi_station_id:
        vi_history = workbook.station_events(vi_station_id).sort_values("test_date", kind="stable")
        vi_rows: list[dict[str, Any]] = []
        for _, product in result.iterrows():
            aoi_time = pd.to_datetime(product.get("aoi_test_date"), errors="coerce")
            candidates = vi_history[vi_history["dmc_raw"].astype(str).eq(str(product["dmc_raw"]))]
            if pd.notna(aoi_time):
                later = candidates[pd.to_datetime(candidates["test_date"], errors="coerce").ge(aoi_time)]
                if not later.empty:
                    candidates = later
            if candidates.empty:
                continue
            event = candidates.iloc[-1]
            failure_values = []
            for name, value in event.items():
                normalized = _norm(name)
                if any(token in normalized for token in ("failurescode", "failurecode", "blockcode")):
                    if pd.notna(value) and _text(value):
                        failure_values.append(_text(value))
            vi_rows.append({
                "dmc_raw": str(product["dmc_raw"]), "vi_event_id": event["event_id"],
                "vi_test_date": event["test_date"], "vi_state": event["state"],
                "vi_state_raw": event["state_raw"], "vi_defect_code": ";".join(failure_values),
            })
        if vi_rows:
            result = result.merge(pd.DataFrame(vi_rows).drop_duplicates("dmc_raw", keep="last"), on="dmc_raw", how="left")
    for column in ("vi_event_id", "vi_test_date", "vi_state", "vi_state_raw", "vi_defect_code"):
        if column not in result:
            result[column] = pd.NA
    result = result.sort_values(["aoi_test_date", "dmc_raw"], na_position="last", kind="stable").reset_index(drop=True)
    result["global_order"] = range(1, len(result) + 1)
    return result


def process_parameter_frame(workbook: StationWorkbookData) -> pd.DataFrame:
    """Return one DMC row with the latest numeric WP1-WP5 parameters."""
    parameters = workbook.parameters.copy()
    parameters = parameters[
        parameters["station_id"].astype(str).str.contains(r"_wp[1-5]$", regex=True)
        & parameters["numeric_value"].notna()
    ]
    if parameters.empty:
        return pd.DataFrame(columns=["dmc_raw"])
    parameters["feature_name"] = (
        parameters["station_id"].astype(str) + "." + parameters["parameter_name"].astype(str)
    )
    parameters = parameters.sort_values("test_date", na_position="last", kind="stable")
    parameters = parameters.drop_duplicates(["dmc_raw", "feature_name"], keep="last")
    result = (
        parameters.pivot(index="dmc_raw", columns="feature_name", values="numeric_value")
        .reset_index().rename_axis(columns=None)
    )
    for column in result.columns[1:]:
        result[column] = pd.to_numeric(result[column], errors="coerce")
    return result
