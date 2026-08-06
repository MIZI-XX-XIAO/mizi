"""本文件读取测试工作簿、解析容差并生成可追溯的Excel质量诊断结果。"""

from __future__ import annotations

from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable
import json
import re

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .analysis_service import CancellationToken
from .data_quality import DataQualityReport

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402


@dataclass(frozen=True)
class ExcelAnalysisRequest:
    workbook_path: Path
    output_parent: Path
    task_name: str
    data_sheet: str = "Data"
    query_sheet: str = "Query parameter"
    column_mapping: dict[str, str] = field(default_factory=dict)
    excel_profile: str = "auto"


@dataclass(frozen=True)
class ParameterSpec:
    name: str
    value_column: str
    tolerance_column: str | None
    lower: float | None
    upper: float | None
    parse_status: str


@dataclass
class ExcelWorkbookData:
    raw_data: pd.DataFrame
    data: pd.DataFrame
    query_parameters: dict[str, Any]
    parameter_specs: list[ParameterSpec]
    quality_report: DataQualityReport
    data_sheet: str
    query_sheet: str | None
    column_mapping: dict[str, str]
    excel_profile: str


@dataclass(frozen=True)
class ExcelProgressEvent:
    stage: str
    percent: int
    message: str


@dataclass
class ExcelAnalysisCallbacks:
    on_progress: Callable[[ExcelProgressEvent], None] = lambda _event: None


@dataclass
class ExcelAnalysisResult:
    status: str
    output_dir: Path
    summary: dict[str, Any]
    frames: dict[str, pd.DataFrame] = field(default_factory=dict)
    workbook_data: ExcelWorkbookData | None = None


CORE_ALIASES = {
    "dmc_raw": ("ident no", "ident no.", "identno", "dmc", "dmc raw", "dmc_raw"),
    "ttnr": ("ttnr",),
    "variant": ("variant",),
    "batch": ("batch", "batch id", "batch_id"),
    "state": ("state", "state test", "result state", "result.aoi result", "aoi result"),
    "line": ("line", "location", "location(s)"),
    "test_date": ("test date", "testdate", "test_date"),
    "result_timestamp": (
        "result.timestamping", "result timestamping", "timestamp", "production timestamp"
    ),
}


def _normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _unique_headers(values: list[Any]) -> tuple[list[str], list[str]]:
    originals = [str(value).strip() if value is not None else "" for value in values]
    counts: Counter[str] = Counter()
    unique: list[str] = []
    for index, original in enumerate(originals, 1):
        base = original or f"Unnamed_{index}"
        counts[base] += 1
        unique.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return unique, originals


def inspect_excel_workbook(path: Path) -> dict[str, list[str]]:
    """快速读取工作表和首行表头，供GUI配置导入映射。"""
    resolved = path.resolve()
    if resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("首版仅支持 .xlsx 和 .xlsm 工作簿")
    if not resolved.is_file():
        raise FileNotFoundError(f"Excel工作簿不存在：{resolved}")
    workbook = load_workbook(
        resolved, read_only=True, data_only=True, keep_vba=resolved.suffix.lower() == ".xlsm"
    )
    try:
        result: dict[str, list[str]] = {}
        for name in workbook.sheetnames:
            row = next(workbook[name].iter_rows(min_row=1, max_row=1, values_only=True), ())
            result[name] = [str(value).strip() if value is not None else "" for value in row]
        return result
    finally:
        workbook.close()


def preview_excel_sheet(path: Path, sheet_name: str, max_rows: int = 50) -> pd.DataFrame:
    """只读取少量数据行用于导入前预览，不触发完整分析。"""
    resolved = path.resolve()
    workbook = load_workbook(
        resolved, read_only=True, data_only=True, keep_vba=resolved.suffix.lower() == ".xlsm"
    )
    try:
        if sheet_name not in workbook.sheetnames:
            return pd.DataFrame()
        rows = workbook[sheet_name].iter_rows(values_only=True)
        first = next(rows, None)
        if first is None:
            return pd.DataFrame()
        columns, _originals = _unique_headers(list(first))
        values = []
        for index, row in enumerate(rows):
            if index >= max_rows:
                break
            values.append(row)
        return pd.DataFrame(values, columns=columns)
    finally:
        workbook.close()


def _read_sheet(workbook, name: str) -> tuple[pd.DataFrame, list[str], list[str]]:
    worksheet = workbook[name]
    rows = worksheet.iter_rows(values_only=True)
    first = next(rows, None)
    if first is None:
        return pd.DataFrame(), [], []
    unique, originals = _unique_headers(list(first))
    values = list(rows)
    return pd.DataFrame(values, columns=unique), unique, originals


def _read_query_parameters(workbook, sheet_name: str | None) -> dict[str, Any]:
    if not sheet_name or sheet_name not in workbook.sheetnames:
        return {}
    result: dict[str, Any] = {}
    for row in workbook[sheet_name].iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if key and _normalized_header(key) != "queryparameter":
            result[key] = row[1] if len(row) > 1 else None
    return result


def _auto_mapping(columns: list[str], originals: list[str]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for unique, original in zip(columns, originals):
        lookup.setdefault(_normalized_header(original), unique)
    mapping: dict[str, str] = {}
    for canonical, aliases in CORE_ALIASES.items():
        for alias in aliases:
            if source := lookup.get(_normalized_header(alias)):
                mapping[canonical] = source
                break
    return mapping


def _parse_tolerance(value: Any) -> tuple[float, float] | None:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    text = str(value).strip().replace("−", "-").replace(",", "")
    if not text:
        return None
    number = r"[-+]?\d+(?:\.\d+)?"
    patterns = (
        rf"^\s*({number})\s*(?:\.{{2,}}|…|~|～|to)\s*({number})\s*$",
        rf"^\s*({number})\s+-\s+({number})\s*$",
    )
    for pattern in patterns:
        match = re.match(pattern, text, flags=re.IGNORECASE)
        if match:
            lower, upper = float(match.group(1)), float(match.group(2))
            return (lower, upper) if lower <= upper else None
    return None


def _parameter_specs(
    frame: pd.DataFrame, columns: list[str], originals: list[str], report: DataQualityReport
) -> list[ParameterSpec]:
    specs: list[ParameterSpec] = []
    used_tolerances: set[str] = set()
    for index, (column, original) in enumerate(zip(columns, originals)):
        if not str(original).strip().lower().startswith("result."):
            continue
        if _normalized_header(original) == "resulttimestamping":
            continue
        if pd.to_numeric(frame[column], errors="coerce").notna().sum() == 0:
            report.warnings.append(f"结果列 {original} 不是数值参数，已保留原始数据但不参与容差统计")
            continue
        tolerance_column: str | None = None
        for next_index in range(index + 1, len(columns)):
            next_original = str(originals[next_index]).strip()
            if next_original.lower().startswith("result."):
                break
            if _normalized_header(next_original).startswith("tolerance"):
                candidate = columns[next_index]
                if candidate not in used_tolerances:
                    tolerance_column = candidate
                    used_tolerances.add(candidate)
                break
        parsed: list[tuple[float, float]] = []
        invalid_nonempty = 0
        if tolerance_column:
            for value in frame[tolerance_column]:
                bounds = _parse_tolerance(value)
                if bounds:
                    parsed.append(bounds)
                elif pd.notna(value) and str(value).strip():
                    invalid_nonempty += 1
        representative = Counter(parsed).most_common(1)[0][0] if parsed else (None, None)
        if not tolerance_column:
            status = "missing_tolerance"
            report.warnings.append(f"参数 {original} 后未找到对应Tolerance列")
        elif not parsed:
            status = "no_valid_tolerance"
            report.warnings.append(f"参数 {original} 没有可解析的Tolerance")
        elif invalid_nonempty:
            status = "partially_invalid"
            report.warnings.append(f"参数 {original} 有 {invalid_nonempty} 个Tolerance无法解析")
        else:
            status = "valid"
        specs.append(ParameterSpec(
            name=str(original), value_column=column, tolerance_column=tolerance_column,
            lower=representative[0], upper=representative[1], parse_status=status,
        ))
    return specs


def _resolve_excel_profile(originals: list[str], requested: str) -> str:
    if requested not in {"auto", "wp", "aoi", "vi"}:
        raise ValueError(f"未知Excel分析档：{requested}")
    if requested != "auto":
        return requested
    normalized = {_normalized_header(value) for value in originals}
    if normalized & {"failuresarea", "failurescode", "stationno", "documentversion"}:
        return "vi"
    if "aoifailurecode" in normalized:
        return "aoi"
    return "wp"


def _categorical_statistics(frame: pd.DataFrame, originals: list[str], profile: str) -> pd.DataFrame:
    wanted = {
        "aoi": {"aoifailurecode"},
        "vi": {
            "blockcode", "documentversion", "fail1", "failuresarea", "failurescode",
            "resultaoiresult", "resultfrom", "stationno",
        },
    }.get(profile, set())
    rows: list[dict[str, Any]] = []
    for column, original in zip(frame.columns, originals):
        if _normalized_header(original) not in wanted:
            continue
        values = frame[column].dropna().astype(str).str.strip()
        values = values[values.ne("")]
        for value, count in values.value_counts(dropna=False).items():
            rows.append({
                "字段": str(original), "值": value, "数量": int(count),
                "占比": float(count / len(values)) if len(values) else np.nan,
            })
    return pd.DataFrame(rows, columns=["字段", "值", "数量", "占比"])


def load_excel_workbook(
    path: Path,
    data_sheet: str = "Data",
    query_sheet: str = "Query parameter",
    column_mapping: dict[str, str] | None = None,
    excel_profile: str = "auto",
) -> ExcelWorkbookData:
    """读取真实工作簿并保留原始列，同时生成统一业务字段和参数定义。"""
    resolved = path.resolve()
    if resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("首版仅支持 .xlsx 和 .xlsm 工作簿")
    if not resolved.is_file():
        raise FileNotFoundError(f"Excel工作簿不存在：{resolved}")
    workbook = load_workbook(
        resolved, read_only=True, data_only=True, keep_vba=resolved.suffix.lower() == ".xlsm"
    )
    try:
        if data_sheet not in workbook.sheetnames:
            raise ValueError(f"找不到数据工作表 {data_sheet}；可用：{', '.join(workbook.sheetnames)}")
        raw, columns, originals = _read_sheet(workbook, data_sheet)
        query_name = query_sheet if query_sheet in workbook.sheetnames else None
        query = _read_query_parameters(workbook, query_name)
    finally:
        workbook.close()

    resolved_profile = _resolve_excel_profile(originals, excel_profile)
    report = DataQualityReport("Excel测试数据", len(raw), len(raw.columns))
    if raw.empty:
        report.errors.append("Data工作表没有数据行")
    automatic = _auto_mapping(columns, originals)
    supplied = column_mapping or {}
    mapping = {**automatic, **{key: value for key, value in supplied.items() if value}}
    for canonical, source in list(mapping.items()):
        if source not in raw.columns:
            report.errors.append(f"字段映射 {canonical} 指向不存在的列：{source}")
            mapping.pop(canonical)
    data = raw.copy()
    for canonical, source in mapping.items():
        data[canonical] = raw[source]
    if "dmc_raw" in data:
        data["dmc_raw"] = data["dmc_raw"].map(
            lambda value: "" if pd.isna(value) else str(value).strip()
        )
    for name in ("state", "variant", "batch", "ttnr", "line"):
        if name in data:
            data[name] = data[name].map(lambda value: "" if pd.isna(value) else str(value).strip())
    if "state" in data:
        data["state"] = data["state"].str.upper()
    for name in ("test_date", "result_timestamp"):
        if name in data:
            converted = pd.to_datetime(data[name], errors="coerce")
            invalid = int(data[name].notna().sum() - converted.notna().sum())
            data[name] = converted
            if invalid:
                report.warnings.append(f"{name} 有 {invalid} 个值无法转换为日期")
    if "result_timestamp" in data:
        data["timestamp"] = data["result_timestamp"]
    elif "test_date" in data:
        data["timestamp"] = data["test_date"]
    if "dmc_raw" not in mapping:
        report.warnings.append("未识别 Ident No.，独立分析可继续，但无法按产品码联合图片结果")
    if "state" not in mapping:
        report.warnings.append("未识别 State，将只按Tolerance计算质量结果")
    duplicates = int(data["dmc_raw"].duplicated().sum()) if "dmc_raw" in data else 0
    if duplicates:
        report.warnings.append(f"Ident No. 有 {duplicates} 个重复值，将保留全部测试记录")
    specs = _parameter_specs(raw, columns, originals, report)
    if not specs and resolved_profile != "vi":
        report.errors.append("未识别任何 Result.* 参数列")
    elif not specs:
        report.warnings.append("VI工作簿没有数值Result.*参数，将执行分类失效项统计")
    report.metrics.update({
        "记录数量": len(data), "字段数量": len(raw.columns), "参数数量": len(specs),
        "可解析容差参数": sum(spec.parse_status in {"valid", "partially_invalid"} for spec in specs),
        "空单元格": int(raw.isna().sum().sum()), "数据工作表": data_sheet,
        "查询参数工作表": query_name or "未找到",
    })
    return ExcelWorkbookData(
        raw_data=raw, data=data, query_parameters=query, parameter_specs=specs,
        quality_report=report, data_sheet=data_sheet, query_sheet=query_name,
        column_mapping=mapping, excel_profile=resolved_profile,
    )


def excel_relationship_frame(
    workbook_data: ExcelWorkbookData,
    analyzed_frame: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """提取联合分析所需的产品键、时间、分组字段和真实测量参数。"""
    source = analyzed_frame if analyzed_frame is not None else workbook_data.data
    wanted = (
        "dmc_raw", "timestamp", "test_date", "result_timestamp",
        "ttnr", "variant", "batch", "line",
    )
    columns = [name for name in wanted if name in source]
    columns.extend(
        spec.value_column for spec in workbook_data.parameter_specs
        if spec.value_column in source and spec.value_column not in columns
    )
    return source.loc[:, columns].copy()


def _state_result(value: Any) -> str:
    normalized = str(value or "").strip().upper()
    if normalized in {"OK", "GOOD", "PASS", "PASSED"}:
        return "OK"
    if normalized in {"NOK", "NG", "FAIL", "FAILED"}:
        return "NOK"
    return "UNKNOWN"


def _group_quality(frame: pd.DataFrame, dimensions: tuple[str, ...]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for dimension in dimensions:
        if dimension not in frame or not frame[dimension].astype(str).str.strip().ne("").any():
            continue
        for value, group in frame.groupby(dimension, dropna=False):
            state_known = group["state_result"].ne("UNKNOWN")
            tolerance_known = group["tolerance_result"].ne("UNKNOWN")
            rows.append({
                "维度": dimension, "值": str(value), "记录数": len(group),
                "原始NOK数": int(group["state_result"].eq("NOK").sum()),
                "原始NOK率": float(group.loc[state_known, "state_result"].eq("NOK").mean()) if state_known.any() else np.nan,
                "超差数": int(group["tolerance_result"].eq("NOK").sum()),
                "超差率": float(group.loc[tolerance_known, "tolerance_result"].eq("NOK").mean()) if tolerance_known.any() else np.nan,
                "判定冲突数": int(group["judgement_conflict"].sum()),
            })
    return pd.DataFrame(rows)


def _write_visualizations(trend: pd.DataFrame, parameter_stats: pd.DataFrame, output_dir: Path) -> None:
    visualizations = output_dir / "visualizations"
    visualizations.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(11, 4.8))
    if not trend.empty:
        ax.plot(trend["date"], trend["state_nok_rate"], marker="o", label="State NOK rate")
        ax.plot(trend["date"], trend["tolerance_nok_rate"], marker="o", label="Tolerance violation rate")
        ax.legend()
    else:
        ax.text(0.5, 0.5, "No valid timestamp for trend", ha="center", transform=ax.transAxes)
    ax.set_ylim(0, 1); ax.set_title("Quality trend"); ax.set_ylabel("rate"); ax.grid(alpha=0.25)
    fig.autofmt_xdate(); fig.tight_layout(); fig.savefig(visualizations / "quality_trend.png", dpi=150); plt.close(fig)

    top = parameter_stats.sort_values("超差数", ascending=False).head(15) if not parameter_stats.empty else parameter_stats
    fig, ax = plt.subplots(figsize=(10, 5.5))
    if not top.empty:
        ax.barh(top["参数"].astype(str), top["超差数"].astype(float), color="#289DDB")
        ax.invert_yaxis()
    else:
        ax.text(0.5, 0.5, "No evaluable parameter", ha="center", transform=ax.transAxes)
    ax.set_title("Top tolerance violations"); ax.set_xlabel("violation count")
    fig.tight_layout(); fig.savefig(visualizations / "top_tolerance_violations.png", dpi=150); plt.close(fig)


def _safe_name(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", value.strip()).rstrip(". ")
    if not cleaned:
        raise ValueError("任务名称不能为空")
    return cleaned[:80]


def analyze_excel_quality(
    request: ExcelAnalysisRequest,
    callbacks: ExcelAnalysisCallbacks | None = None,
    token: CancellationToken | None = None,
) -> ExcelAnalysisResult:
    """独立执行Excel质量诊断并输出可复核的标准化表和统计结果。"""
    callbacks = callbacks or ExcelAnalysisCallbacks()
    token = token or CancellationToken()
    callbacks.on_progress(ExcelProgressEvent("LOADING", 5, "读取工作簿"))
    workbook_data = load_excel_workbook(
        request.workbook_path, request.data_sheet, request.query_sheet,
        request.column_mapping, request.excel_profile,
    )
    if workbook_data.quality_report.errors:
        raise ValueError("Excel数据检查失败：" + "；".join(workbook_data.quality_report.errors))
    if token.is_cancelled:
        return ExcelAnalysisResult("cancelled", Path(), {"status": "cancelled"}, workbook_data=workbook_data)

    frame = workbook_data.data.copy()
    frame.insert(0, "source_row", np.arange(2, len(frame) + 2))
    frame["state_result"] = frame["state"].map(_state_result) if "state" in frame else "UNKNOWN"
    violation_rows: list[dict[str, Any]] = []
    evaluated = np.zeros(len(frame), dtype=int)
    violations = np.zeros(len(frame), dtype=int)
    stats: list[dict[str, Any]] = []
    total_specs = max(1, len(workbook_data.parameter_specs))
    for index, spec in enumerate(workbook_data.parameter_specs, 1):
        if token.is_cancelled:
            return ExcelAnalysisResult("cancelled", Path(), {"status": "cancelled"}, workbook_data=workbook_data)
        values = pd.to_numeric(frame[spec.value_column], errors="coerce")
        valid_value = values.notna()
        parameter_evaluated = 0
        parameter_violations = 0
        bounds_seen: list[tuple[float, float]] = []
        for position in range(len(frame)):
            if not valid_value.iloc[position] or not spec.tolerance_column:
                continue
            bounds = _parse_tolerance(frame.iloc[position][spec.tolerance_column])
            if not bounds:
                continue
            lower, upper = bounds
            value = float(values.iloc[position])
            evaluated[position] += 1
            parameter_evaluated += 1
            bounds_seen.append(bounds)
            if value < lower or value > upper:
                violations[position] += 1
                parameter_violations += 1
                violation_rows.append({
                    "source_row": int(frame.iloc[position]["source_row"]),
                    "dmc_raw": frame.iloc[position].get("dmc_raw", ""),
                    "参数": spec.name, "值": value, "下限": lower, "上限": upper,
                    "Tolerance原值": frame.iloc[position][spec.tolerance_column],
                })
        common = Counter(bounds_seen).most_common(1)[0][0] if bounds_seen else (np.nan, np.nan)
        stats.append({
            "参数": spec.name, "有效数值": int(valid_value.sum()),
            "缺失率": round(float(1 - valid_value.mean()), 5) if len(frame) else np.nan,
            "最小值": float(values.min()) if valid_value.any() else np.nan,
            "最大值": float(values.max()) if valid_value.any() else np.nan,
            "均值": float(values.mean()) if valid_value.any() else np.nan,
            "常用下限": common[0], "常用上限": common[1], "可判定数": parameter_evaluated,
            "超差数": parameter_violations,
            "超差率": parameter_violations / parameter_evaluated if parameter_evaluated else np.nan,
            "容差状态": spec.parse_status,
        })
        callbacks.on_progress(ExcelProgressEvent(
            "ANALYZING", 15 + round(index / total_specs * 55), f"分析参数 {index}/{total_specs}"
        ))

    frame["evaluated_parameter_count"] = evaluated
    frame["tolerance_violation_count"] = violations
    frame["tolerance_result"] = np.where(evaluated == 0, "UNKNOWN", np.where(violations > 0, "NOK", "OK"))
    frame["judgement_conflict"] = (
        frame["state_result"].ne("UNKNOWN") & frame["tolerance_result"].ne("UNKNOWN")
        & frame["state_result"].ne(frame["tolerance_result"])
    )
    violation_frame = pd.DataFrame(violation_rows, columns=[
        "source_row", "dmc_raw", "参数", "值", "下限", "上限", "Tolerance原值"
    ])
    conflict_frame = frame[frame["judgement_conflict"]].copy()
    parameter_stat_columns = [
        "参数", "有效数值", "缺失率", "最小值", "最大值", "均值", "常用下限", "常用上限",
        "可判定数", "超差数", "超差率", "容差状态",
    ]
    parameter_stats = pd.DataFrame(stats, columns=parameter_stat_columns)
    if not parameter_stats.empty:
        parameter_stats = parameter_stats.sort_values("超差数", ascending=False, ignore_index=True)
    categorical_stats = _categorical_statistics(
        workbook_data.raw_data, list(workbook_data.raw_data.columns), workbook_data.excel_profile
    )
    group_stats = _group_quality(frame, ("variant", "batch", "line", "ttnr"))
    time_column = next((name for name in ("result_timestamp", "test_date") if name in frame), None)
    trend = pd.DataFrame(columns=["date", "records", "state_nok_rate", "tolerance_nok_rate"])
    if time_column:
        dated = frame.dropna(subset=[time_column]).copy()
        if not dated.empty:
            dated["date"] = pd.to_datetime(dated[time_column]).dt.date
            trend = dated.groupby("date", as_index=False).agg(
                records=("source_row", "size"),
                state_nok_rate=("state_result", lambda values: float(values.eq("NOK").mean())),
                tolerance_nok_rate=("tolerance_result", lambda values: float(values.eq("NOK").mean())),
            )
    known_state = frame["state_result"].ne("UNKNOWN")
    known_tolerance = frame["tolerance_result"].ne("UNKNOWN")
    date_values = pd.to_datetime(frame[time_column], errors="coerce") if time_column else pd.Series(dtype="datetime64[ns]")
    summary = {
        "status": "complete", "workbook": str(request.workbook_path.resolve()),
        "data_sheet": workbook_data.data_sheet, "query_sheet": workbook_data.query_sheet,
        "excel_profile": workbook_data.excel_profile,
        "record_count": len(frame), "parameter_count": len(workbook_data.parameter_specs),
        "state_ok_count": int(frame["state_result"].eq("OK").sum()),
        "state_nok_count": int(frame["state_result"].eq("NOK").sum()),
        "state_nok_rate": float(frame.loc[known_state, "state_result"].eq("NOK").mean()) if known_state.any() else None,
        "tolerance_nok_count": int(frame["tolerance_result"].eq("NOK").sum()),
        "tolerance_nok_rate": float(frame.loc[known_tolerance, "tolerance_result"].eq("NOK").mean()) if known_tolerance.any() else None,
        "violation_event_count": len(violation_frame),
        "judgement_conflict_count": int(frame["judgement_conflict"].sum()),
        "date_start": date_values.min().isoformat() if not date_values.empty and pd.notna(date_values.min()) else None,
        "date_end": date_values.max().isoformat() if not date_values.empty and pd.notna(date_values.max()) else None,
        "warning": "State为原始结论；Tolerance结果为系统重算，两者不一致时需人工复核。",
    }
    callbacks.on_progress(ExcelProgressEvent("WRITING", 78, "写入诊断结果"))
    output_parent = request.output_parent.resolve()
    output_parent.mkdir(parents=True, exist_ok=True)
    base = f"{_safe_name(request.task_name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    output_dir = output_parent / base
    sequence = 1
    while output_dir.exists():
        output_dir = output_parent / f"{base}_{sequence:02d}"; sequence += 1
    output_dir.mkdir(parents=False)
    frame.to_csv(output_dir / "excel_standardized_data.csv", index=False, encoding="utf-8-sig")
    parameter_stats.to_csv(output_dir / "excel_parameter_statistics.csv", index=False, encoding="utf-8-sig")
    categorical_stats.to_csv(output_dir / "excel_categorical_statistics.csv", index=False, encoding="utf-8-sig")
    violation_frame.to_csv(output_dir / "excel_tolerance_violations.csv", index=False, encoding="utf-8-sig")
    conflict_frame.to_csv(output_dir / "excel_judgement_conflicts.csv", index=False, encoding="utf-8-sig")
    group_stats.to_csv(output_dir / "excel_group_quality.csv", index=False, encoding="utf-8-sig")
    trend.to_csv(output_dir / "excel_quality_trend.csv", index=False, encoding="utf-8-sig")
    workbook_data.quality_report.to_frame().to_csv(
        output_dir / "excel_data_quality.csv", index=False, encoding="utf-8-sig"
    )
    (output_dir / "excel_analysis_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output_dir / "excel_parameter_specs.json").write_text(
        json.dumps([asdict(spec) for spec in workbook_data.parameter_specs], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "excel_query_parameters.json").write_text(
        json.dumps(workbook_data.query_parameters, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    callbacks.on_progress(ExcelProgressEvent("VISUALIZING", 90, "生成质量趋势图"))
    _write_visualizations(trend, parameter_stats, output_dir)
    callbacks.on_progress(ExcelProgressEvent("COMPLETE", 100, "Excel分析完成"))
    return ExcelAnalysisResult("complete", output_dir, summary, {
        "standardized": frame, "parameter_stats": parameter_stats, "violations": violation_frame,
        "conflicts": conflict_frame, "group_stats": group_stats, "trend": trend,
        "quality": workbook_data.quality_report.to_frame(), "categorical_stats": categorical_stats,
    }, workbook_data)
