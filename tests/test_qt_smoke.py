"""本文件使用pytest-qt离屏验证主窗口、参数模型和后台Worker能够创建及安全关闭。"""

import os
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
pytest.importorskip("PySide6")
from PySide6.QtCore import QSize, Qt  # noqa: E402
from PySide6.QtWidgets import QLabel  # noqa: E402

from gui.main_window import MainWindow  # noqa: E402
from gui.parameter_dialog import ParameterDialog  # noqa: E402
from gui.workbench import LayoutProfile, WorkbenchStack, resolve_layout_profile  # noqa: E402
from src.result_views import ResultView  # noqa: E402


def test_main_window_starts(qtbot) -> None:
    window = MainWindow(Path(__file__).resolve().parents[1])
    qtbot.addWidget(window)
    window.show()
    assert window.windowTitle() == "MEA多工站缺陷规律分析"
    assert window.tabs.count() == 7
    assert isinstance(window.tabs, WorkbenchStack)
    assert "Excel分析" in window.tabs.tabText(4)
    assert "关联分析" in window.tabs.tabText(5)
    assert window.workbench.navigation.buttons[0].isChecked()
    assert window.workbench.assistant.connection.text() == "未连接"
    assert "#091424" in window.styleSheet()
    assert window.minimumWidth() <= 980
    window.close()
    window.deleteLater()


def test_result_cards_open_filtered_dialogs_and_alert_colors_are_readable(qtbot) -> None:
    window = MainWindow(Path(__file__).resolve().parents[1])
    qtbot.addWidget(window)
    window.show()
    sections = {
        "periodic": pd.DataFrame(columns=["pattern_id"]),
        "burst": pd.DataFrame(columns=["pattern_id"]),
        "code": pd.DataFrame([{"canonical_code": "5520", "pattern_type": "periodic"}]),
        "trajectory": pd.DataFrame(columns=["trajectory_id"]),
        "cooccurrence": pd.DataFrame(columns=["缺陷A", "缺陷B"]),
        "transition": pd.DataFrame(columns=["前一缺陷", "后一缺陷"]),
        "other": pd.DataFrame(columns=["pattern_id"]),
    }
    alerts = pd.DataFrame([{"severity": "warning", "alert_at_order": 2}])
    window._current_result_view = ResultView(
        pd.DataFrame(), pd.DataFrame(), alerts, sections, {},
    )
    window.evidence_mode.setCurrentIndex(window.evidence_mode.findData("code"))

    qtbot.mouseClick(window.pattern_result_card, Qt.LeftButton)
    assert window.pattern_dialog.isVisible()
    assert window.pattern_dialog.tabs.currentWidget() is window.pattern_dialog.widgets["code"]
    assert "(1)" in window.pattern_dialog.tabs.tabText(
        window.pattern_dialog.tabs.indexOf(window.pattern_dialog.widgets["code"])
    )

    qtbot.mouseClick(window.alert_result_card, Qt.LeftButton)
    assert window.alert_dialog.isVisible()
    model = window.alert_dialog.table.model
    index = model.index(0, 0)
    assert model.data(index, Qt.BackgroundRole).name() == "#493b20"
    assert model.data(index, Qt.ForegroundRole).name() == "#ffe09b"
    window.pattern_dialog.close()
    window.alert_dialog.close()
    window.close()
    window.deleteLater()


def test_parameter_dialog_keeps_independent_detection_profiles(qtbot) -> None:
    project = Path(__file__).resolve().parents[1]
    import yaml
    config = yaml.safe_load((project / "config/analysis_config.yaml").read_text(encoding="utf-8"))
    dialog = ParameterDialog(config, project / "config/analysis_config.yaml")
    qtbot.addWidget(dialog)
    dialog.editors["red_min"].setValue(111)
    dialog.profile_combo.setCurrentText("5X")
    dialog.editors["red_min"].setValue(222)
    values = dialog.values()
    assert values["detection_profiles"]["5S"]["red_min"] == 111
    assert values["detection_profiles"]["5X"]["red_min"] == 222
    assert values["detection_profiles"]["7S"]["red_min"] == 150
    dialog.close()


def test_workbench_navigation_and_responsive_layout(qtbot) -> None:
    window = MainWindow(Path(__file__).resolve().parents[1])
    qtbot.addWidget(window)
    window.resize(1600, 900)
    window.show()

    window.workbench.navigation.buttons[5].click()
    assert window.tabs.currentIndex() == 5
    assert window.workbench.header.title.text() == "关联分析"

    window.workbench.apply_responsive_layout(QSize(1366, 768))
    assert window.workbench.profile is LayoutProfile.COMPACT
    assert window.workbench.navigation.width() == 76
    assert not window.workbench.assistant.isVisible()

    qtbot.wait(10)
    content_width = window.workbench.content.width()
    window.workbench.set_assistant_visible(True)
    qtbot.wait(10)
    assert window.workbench._assistant_overlay
    assert window.workbench.content.width() == content_width
    assert 320 <= window.workbench.assistant.width() <= 400

    window.workbench.apply_responsive_layout(QSize(1920, 1200))
    assert window.workbench.profile is LayoutProfile.FULL
    assert window.workbench.navigation.width() == 214
    assert window.workbench.assistant.isVisible()
    window.close()
    window.deleteLater()


def test_layout_profiles_cover_windows_scaling_targets() -> None:
    assert resolve_layout_profile(QSize(1920, 1200)) is LayoutProfile.FULL
    assert resolve_layout_profile(QSize(1536, 960)) is LayoutProfile.FULL
    assert resolve_layout_profile(QSize(1280, 800)) is LayoutProfile.COMPACT
    assert resolve_layout_profile(QSize(1366, 768)) is LayoutProfile.COMPACT
    assert resolve_layout_profile(QSize(980, 620)) is LayoutProfile.TIGHT


def test_compact_excel_page_does_not_set_oversized_window_hint(qtbot) -> None:
    window = MainWindow(Path(__file__).resolve().parents[1])
    qtbot.addWidget(window)
    window.resize(1280, 800)
    window.show()
    window.tabs.setCurrentIndex(4)
    window.workbench.apply_responsive_layout(QSize(1280, 800), force=True)
    qtbot.wait(10)

    assert window.minimumSizeHint().height() < 800
    assert window.workbench.content.width() >= 1100
    assert window.excel_page.result_tabs.usesScrollButtons()
    assert window.excel_page.config_scroll.horizontalScrollBar().maximum() == 0
    assert not window.workbench.header.subtitle.isVisible()
    assert not window.workbench.header.task_context.isVisible()
    window.close()
    window.deleteLater()


def test_assistant_shell_is_offline_and_context_is_explicit(qtbot) -> None:
    window = MainWindow(Path(__file__).resolve().parents[1])
    qtbot.addWidget(window)
    window.show()
    assistant = window.workbench.assistant

    assert not assistant.backend.available
    assert assistant.connection.text() == "未连接"
    assert "不会发送" in assistant.findChild(QLabel, "privacyBanner").text()

    window.tabs.setCurrentIndex(6)
    assert "图片复核" in assistant.context_label.text()
    assistant.clear_context()
    assert "尚未附加" in assistant.context_label.text()
    window.close()
    window.deleteLater()


def test_pattern_evidence_strip_and_fullscreen_review(qtbot, tmp_path: Path) -> None:
    project = Path(__file__).resolve().parents[1]
    import yaml
    config = yaml.safe_load((project / "config/analysis_config.yaml").read_text(encoding="utf-8"))
    a_image = np.full((80, 120), 70, dtype=np.uint8)
    e_image = cv2.cvtColor(a_image, cv2.COLOR_GRAY2BGR)
    cv2.rectangle(e_image, (30, 20), (42, 32), (0, 0, 255), thickness=-1)
    a_path, e_path = tmp_path / "a.png", tmp_path / "e.png"
    cv2.imencode(".png", a_image)[1].tofile(str(a_path))
    cv2.imencode(".png", e_image)[1].tofile(str(e_path))
    products = pd.DataFrame({
        "global_order": [1, 2, 3], "analysis_scope": ["5S"] * 3,
        "order_code": ["P1", "P2", "P3"], "dmc_raw": ["P1", "P2", "P3"],
        "camera": ["5S"] * 3, "a_image_path": [str(a_path)] * 3,
        "e_image_path": [str(e_path)] * 3,
    })
    window = MainWindow(project)
    qtbot.addWidget(window)
    window.show()
    window.tabs.setCurrentWidget(window.review)
    window.review.set_data(products, pd.DataFrame({"global_order": pd.Series(dtype=int)}), config)
    qtbot.waitUntil(lambda: window.review._payload is not None, timeout=10_000)

    pattern = pd.Series({
        "pattern_id": "5S-P001", "pattern_type": "periodic", "cluster_id": "5S-C001",
        "period": 2, "confidence": 0.9, "first_order": 1,
        "observed_orders": "1;3", "inferred_missing_orders": "2",
    })
    window.tabs.setCurrentIndex(3)
    window._jump_from_pattern(pattern)
    assert window.tabs.currentWidget() is window.review
    assert window.review.pattern_record is not None
    assert window.review.pattern_panel.isVisible()
    assert window.review.pattern_list.count() == 3
    assert window.review.pattern_list.item(1).data(Qt.UserRole + 1) == "missing"
    window.review.pattern_list.setCurrentRow(2)
    qtbot.waitUntil(lambda: window.review.current_order == 3, timeout=5_000)

    window.review._open_pattern_item_fullscreen(window.review.pattern_list.item(0))
    qtbot.waitUntil(
        lambda: window.review._fullscreen is not None and window.review._fullscreen.isVisible(),
        timeout=10_000,
    )
    assert window.review._fullscreen.image_type == "E图"
    assert window.review.current_order == 1
    qtbot.keyClick(window.review._fullscreen, Qt.Key_Escape)
    qtbot.waitUntil(lambda: not window.review._fullscreen.isVisible(), timeout=5_000)
    assert window.review.pattern_panel.isVisible()
    assert window.review.current_order == 1

    code_pattern = pd.Series({
        "pattern_id": "CP0001", "pattern_type": "periodic", "canonical_code": "5011",
        "period": 6, "confidence": 0.82, "first_production_order": 100,
        "observed_production_orders": "100;106", "evidence_task_orders": "1;3",
        "missing_task_orders": "2",
    })
    window._jump_from_pattern(code_pattern)
    assert window.review.current_order == 1
    assert window.review.pattern_list.count() == 3

    joint_evidence = pd.Series({
        "pattern_type": "cluster", "canonical_code": "5011", "spatial_id": "5S-C001",
        "support_task_orders": "2;3",
    })
    window._jump_from_pattern(joint_evidence)
    assert window.review.current_order == 2
    assert window.review.pattern_list.count() == 2

    trajectory = pd.Series({
        "trajectory_id": "5S-T001", "pattern_type": "linear_drift", "task_orders": "1;2;3",
    })
    window._jump_from_pattern(trajectory)
    assert window.review.current_order == 1

    conflict = pd.Series({"global_order": 3, "comparison_status": "label_conflict"})
    window._jump_from_pattern(conflict)
    assert window.review.current_order == 3
    window.close()
    window.deleteLater()


def test_excel_page_runs_analysis_in_background(qtbot, tmp_path: Path) -> None:
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Data"
    sheet.append(["Ident No.", "State", "Result.Force", "Tolerance"])
    sheet.append(["DMC-1", "OK", 30, "25 ... 70"])
    sheet.append(["DMC-2", "NOK", 80, "25 ... 70"])
    path = tmp_path / "gui_excel.xlsx"; workbook.save(path)

    window = MainWindow(Path(__file__).resolve().parents[1])
    qtbot.addWidget(window); window.show(); window.tabs.setCurrentIndex(4)
    page = window.excel_page
    page.workbook_edit.setText(str(path)); page.output_edit.setText(str(tmp_path))
    page.task_name.setText("GUI Excel测试"); page.start_analysis()
    qtbot.waitUntil(lambda: page.current_result is not None, timeout=15_000)
    qtbot.waitUntil(lambda: page.thread is None, timeout=5_000)
    assert page.current_result.summary["tolerance_nok_count"] == 1
    assert window.current_excel_result is page.current_result
    assert window.use_current_excel.isEnabled()
    window.close(); window.deleteLater()


def test_station_task_builds_products_without_company_csv(qtbot, tmp_path: Path) -> None:
    dmc = "376W020BGO57424F00VF004AK"
    workbook = Workbook(); data = workbook.active; data.title = "Data"
    data.append(["Ident No.", "State", "Result.Force", "Tolerance"])
    data.append([dmc, "OK", 30, "25 ... 70"])
    query = workbook.create_sheet("Query parameter")
    query.append(["Query parameter", None]); query.append(["Location(s)", "3003.10.1.1.6"])
    excel_path = tmp_path / "station.xlsx"; workbook.save(excel_path)
    for code in ("DA", "DE"):
        (tmp_path / f"{dmc}20250624{code}.png").write_bytes(b"index-only")

    window = MainWindow(Path(__file__).resolve().parents[1])
    qtbot.addWidget(window)
    window.products_edit.clear()
    window.station_combo.setCurrentIndex(window.station_combo.findData("35_5s_aoi"))
    window.source_excel_edit.setText(str(excel_path))
    window.image_root_edit.setText(str(tmp_path))
    assert window._inspect_products()
    assert len(window.loaded_products) == 1
    assert len(window.analysis_products) == 1
    assert window.analysis_products.iloc[0].a_image_path.endswith("DA.png")
    assert window.analysis_products.iloc[0].e_image_path.endswith("DE.png")
    window.close(); window.deleteLater()
