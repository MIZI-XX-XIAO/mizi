"""本文件测试动态全工站工作簿导入、工艺参数保留和AOI真值顺序。"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from src.station_sources import load_station_catalog
from src.station_workbook import (
    enrich_products_with_station_truth, load_station_workbook, process_parameter_frame,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _catalog():
    return load_station_catalog(PROJECT_ROOT / "config" / "stations.yaml")


def test_flat_all_station_sheet_reads_dynamic_parameters_and_history(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MS03210all"
    sheet.append([
        "Ident No.", "Test Date", "State Test", "Line", "ST", "SI", "FU", "WP", "TP",
        "Result.RollingSpeed", "Tolerance ", "Result.Heatingplatetemperature", "Tolerance  ",
    ])
    sheet.append(["DMC-1", "2026-07-03 10:00:00", "OK", 3003, 10, 1, 1, 1, 1, 3.4, "0 ... 6", None, None])
    sheet.append(["DMC-1", "2026-07-03 10:05:00", "NOK", 3003, 10, 1, 1, 3, 1, 3.8, "0 ... 6", 85, "70 ... 90"])
    path = tmp_path / "all.xlsx"
    workbook.save(path)

    result = load_station_workbook(path, _catalog())

    assert len(result.products) == 1
    assert result.events.station_id.tolist() == ["35_wp1", "35_wp3"]
    assert set(result.parameters.parameter_name) == {
        "Result.RollingSpeed", "Result.Heatingplatetemperature"
    }
    temperature = result.parameters[result.parameters.parameter_name.str.contains("Heating")].iloc[0]
    assert float(temperature.numeric_value) == 85
    assert temperature.tolerance_raw == "70 ... 90"
    relationship = process_parameter_frame(result)
    assert "35_wp1.Result.RollingSpeed" in relationship
    assert "35_wp3.Result.Heatingplatetemperature" in relationship


def test_stacked_station_blocks_are_detected_by_repeated_ident_header(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["3-5 WP1"])
    sheet.append(["Ident No.", "State", "Line", "ST", "SI", "FU", "WP", "TP", "Test Date", "Result.RollingSpeed"])
    sheet.append(["DMC-1", "OK", 3003, 10, 1, 1, 1, 1, "2026-07-03 10:00:00", 3.4])
    sheet.append(["3-5 WP2"])
    sheet.append(["Ident No.", "State", "Line", "ST", "SI", "FU", "WP", "TP", "Test Date", "Result.StaticElectricity"])
    sheet.append(["DMC-1", "OK", 3003, 10, 1, 1, 2, 1, "2026-07-03 10:02:00", -195])
    path = tmp_path / "blocks.xlsx"
    workbook.save(path)

    result = load_station_workbook(path, _catalog())

    assert result.events.station_id.tolist() == ["35_wp1", "35_wp2"]
    assert set(result.parameters.parameter_name) == {
        "Result.RollingSpeed", "Result.StaticElectricity"
    }


def test_aoi_truth_uses_same_day_latest_and_real_timestamp_order(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Ident No.", "State", "Line", "ST", "SI", "FU", "WP", "TP", "Test Date", "Result.AOIFailureCode"])
    sheet.append(["DMC-B", "OK", 3003, 10, 1, 1, 6, 1, "2026-07-03 09:00:00", "1000"])
    sheet.append(["DMC-A", "NOK", 3003, 10, 1, 1, 6, 1, "2026-07-03 10:00:00", "5020"])
    sheet.append(["DMC-A", "OK", 3003, 10, 1, 1, 6, 1, "2026-07-04 10:00:00", "1000"])
    path = tmp_path / "truth.xlsx"
    workbook.save(path)
    station_book = load_station_workbook(path, _catalog())
    products = pd.DataFrame({
        "global_order": [1, 2], "dmc_raw": ["DMC-A", "DMC-B"],
        "capture_date": pd.to_datetime(["2026-07-03", "2026-07-03"]),
    })

    enriched = enrich_products_with_station_truth(products, station_book, "35_5s_aoi")

    assert enriched.dmc_raw.tolist() == ["DMC-B", "DMC-A"]
    assert enriched.aoi_state.tolist() == ["OK", "NOK"]
    assert enriched.global_order.tolist() == [1, 2]
